import openpyxl as op
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font, numbers
from openpyxl.utils import get_column_letter
from src.utils.resource_path import resource_path


def generar_excel(razon_social):
    wb = op.load_workbook(resource_path("./results/"+razon_social+".xlsx"))
    ws = wb.active
    
    fill = PatternFill(start_color="16365C", end_color="16365C", fill_type="solid")
    font_header = Font(name="Arial", size=10, color="FFFFFF", bold=True)
    font_cells = Font(name="Arial", size=10)
    border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    alignment = Alignment(horizontal="center", vertical="center")
    
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = alignment
            cell.font = font_cells
            if cell.row == 1:  # Si la celda está en la primera fila (encabezado)
                cell.fill = fill
                cell.font = font_header
            if cell.column == 4:  # Si la celda está en la cuarta columna
                cell.number_format = "dd/mm/yyyy"
            if cell.column == 8 and cell.row > 1:
                cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                cell.alignment = Alignment(horizontal="right", vertical="center")
    
    column_widths = [8, 13, 17, 13, 4, 8, 8, 10]
    for i, column_width in enumerate(column_widths):
        ws.column_dimensions[get_column_letter(i+1)].width = column_width
    
    last_row = ws.max_row
    # Calcular la suma de todos los valores en esa columna (excluyendo la cabecera)
    column_sum = sum(cell.value for cell in ws['H'][1:last_row] if isinstance(cell.value, (int, float)))
    cell_sum = ws.cell(row=last_row + 1, column=8, value=column_sum)
    cell_sum.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
    cell_sum.alignment = Alignment(horizontal="right", vertical="center")
    cell_sum.font = Font(name="Arial", size=10, bold=True)
    cell_sum.border = border
    
    wb.save(resource_path("./results/"+razon_social+".xlsx"))